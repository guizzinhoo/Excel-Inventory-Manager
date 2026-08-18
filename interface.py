# interface.py
# Front-end CustomTkinter — importa o back-end de Backend.py
# Os dois arquivos e o AvonNatura.xlsx devem ficar na mesma pasta.

import customtkinter as ctk
from tkinter import ttk, messagebox
import pandas as pd
from openpyxl import load_workbook

# ── importa o back-end sem alteração ─────────────────────────────────────────
import Backend as backend

# ── tema ──────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("green")

COR_ROSA  = "#217346"
COR_ROSA2 = "#185C37"
COR_VERDE  = "#2E7D32"
COR_FUNDO  = "#FFF0F5"
COR_CARD   = "#FFFFFF"
COR_TEXTO  = "#212121"
COR_CINZA  = "#757575"
COR_ALERTA = "#FF6F00"

# ── wrappers: chamam a lógica interna do back-end sem usar input() ────────────
# (o Backend.py não é tocado; só reutilizamos a lógica que já existe nele)

def _adicionar_produto(nome, qtd, vendida, valor):
    """Replica adicionar_produto() sem input()."""
    wb = load_workbook(backend.ARQUIVO_EXCEL)
    sheet = wb[backend.ABA_ESTOQUE]
    linha = sheet.max_row
    while sheet.cell(row=linha, column=1).value is None:
        linha -= 1
    linha += 1
    for col, dado in enumerate([nome, int(qtd), int(vendida), None, float(valor), None], start=1):
        sheet.cell(row=linha, column=col, value=dado)
    sheet[f'D{linha}'] = f'=B{linha}-C{linha}'
    sheet[f'F{linha}'] = f'=C{linha}*E{linha}'
    wb.save(backend.ARQUIVO_EXCEL)
    wb.close()

def _registrar_venda(nome, qtd_vender):
    """Replica registrar_venda() sem input(). Retorna (ok, mensagem)."""
    wb = load_workbook(backend.ARQUIVO_EXCEL)
    sheet = wb[backend.ABA_ESTOQUE]
    for linha in range(2, sheet.max_row + 1):
        celula = sheet[f'A{linha}'].value
        if celula is None:
            continue
        if str(celula).strip().lower() == nome.lower():
            qtd_total      = sheet[f'B{linha}'].value or 0
            qtd_ja_vendida = sheet[f'C{linha}'].value or 0
            valor_unitario = sheet[f'E{linha}'].value or 0
            disponivel     = qtd_total - qtd_ja_vendida
            if qtd_vender > disponivel:
                wb.close()
                return False, f'Estoque insuficiente. Disponível: {disponivel}'
            sheet[f'C{linha}'] = qtd_ja_vendida + qtd_vender
            backend.historico_vendas(wb, celula, qtd_vender, valor_unitario)
            wb.save(backend.ARQUIVO_EXCEL)
            wb.close()
            return True, f'Venda de {qtd_vender}x "{celula}" registrada!'
    wb.close()
    return False, 'Produto não encontrado.'

def _atualizar_estoque(nome, qtd_nova):
    """Replica atualizar_estoque() sem input(). Retorna (ok, mensagem)."""
    wb = load_workbook(backend.ARQUIVO_EXCEL)
    sheet = wb[backend.ABA_ESTOQUE]
    for linha in range(2, sheet.max_row + 1):
        celula = sheet[f'A{linha}'].value
        if celula and str(celula).strip().lower() == nome.lower():
            sheet[f'B{linha}'] = (sheet[f'B{linha}'].value or 0) + int(qtd_nova)
            wb.save(backend.ARQUIVO_EXCEL)
            wb.close()
            return True, 'Estoque atualizado com sucesso!'
    wb.close()
    return False, 'Produto não encontrado. Use "Novo Produto" para cadastrá-lo.'

# ── janela principal ──────────────────────────────────────────────────────────

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title('Avon & Natura — Gestão de Estoque')
        self.geometry('1080x680')
        self.minsize(900, 600)
        self.configure(fg_color=COR_FUNDO)
        self._build_layout()
        self._ir_para('estoque')

    # ── estrutura base ────────────────────────────────────────────────────────

    def _build_layout(self):
        # sidebar
        self.sidebar = ctk.CTkFrame(self, width=200, corner_radius=0, fg_color=COR_ROSA)
        self.sidebar.pack(side='left', fill='y')
        self.sidebar.pack_propagate(False)

        ctk.CTkLabel(self.sidebar, text='💄 Avon\n& Natura',
                     font=ctk.CTkFont(size=20, weight='bold'),
                     text_color='white').pack(pady=(30, 30))

        nav = [
            ('📦  Estoque',        'estoque'),
            ('🛒  Registrar Venda', 'venda'),
            ('➕  Novo Produto',   'adicionar'),
            ('🔄  Repor Estoque',  'repor'),
            ('📊  Relatório',      'relatorio'),
            ('📋  Histórico',      'historico'),
            ('⚠️  Estoque Baixo',  'baixo'),
        ]
        for texto, chave in nav:
            ctk.CTkButton(
                self.sidebar, text=texto, anchor='w',
                fg_color='transparent', hover_color=COR_ROSA2,
                text_color='white', font=ctk.CTkFont(size=13),
                height=42, corner_radius=0,
                command=lambda c=chave: self._ir_para(c)
            ).pack(fill='x')

        # área principal
        self.main = ctk.CTkFrame(self, corner_radius=0, fg_color=COR_FUNDO)
        self.main.pack(side='left', fill='both', expand=True)

    def _ir_para(self, chave):
        for w in self.main.winfo_children():
            w.destroy()
        {
            'estoque':   self._tela_estoque,
            'venda':     self._tela_venda,
            'adicionar': self._tela_adicionar,
            'repor':     self._tela_repor,
            'relatorio': self._tela_relatorio,
            'historico': self._tela_historico,
            'baixo':     self._tela_baixo,
        }[chave]()

    # ── helpers de UI ─────────────────────────────────────────────────────────

    def _titulo(self, texto, sub=''):
        ctk.CTkLabel(self.main, text=texto,
                     font=ctk.CTkFont(size=22, weight='bold'),
                     text_color=COR_ROSA).pack(anchor='w', padx=30, pady=(24, 0))
        if sub:
            ctk.CTkLabel(self.main, text=sub,
                         font=ctk.CTkFont(size=13),
                         text_color=COR_CINZA).pack(anchor='w', padx=30)

    def _campo(self, parent, label, placeholder=''):
        ctk.CTkLabel(parent, text=label,
                     font=ctk.CTkFont(size=13, weight='bold'),
                     text_color=COR_TEXTO).pack(anchor='w', pady=(10, 2))
        e = ctk.CTkEntry(parent, placeholder_text=placeholder,
                         height=38, corner_radius=8,
                         border_color='#DDDDDD', font=ctk.CTkFont(size=13))
        e.pack(fill='x')
        return e

    def _botao(self, parent, texto, cmd):
        return ctk.CTkButton(
            parent, text=texto, command=cmd,
            fg_color=COR_ROSA, hover_color=COR_ROSA2,
            height=42, corner_radius=10,
            font=ctk.CTkFont(size=14, weight='bold')
        )

    def _aviso(self, parent, texto, cor=COR_ALERTA):
        ctk.CTkLabel(parent, text=texto,
                     font=ctk.CTkFont(size=12),
                     text_color=cor, wraplength=440).pack(pady=(6, 0))

    def _tabela(self, colunas, linhas):
        frame = ctk.CTkFrame(self.main, fg_color=COR_CARD, corner_radius=12)
        frame.pack(fill='both', expand=True, padx=30, pady=10)

        s = ttk.Style()
        s.theme_use('clam')
        s.configure('T.Treeview',
                    background=COR_CARD, foreground=COR_TEXTO,
                    fieldbackground=COR_CARD, rowheight=28,
                    font=('Segoe UI', 11))
        s.configure('T.Treeview.Heading',
                    background=COR_ROSA, foreground='white',
                    font=('Segoe UI', 11, 'bold'), relief='flat')
        s.map('T.Treeview', background=[('selected', '#FCE4EC')])

        tree = ttk.Treeview(frame, columns=colunas, show='headings',
                            style='T.Treeview', height=18)
        for col in colunas:
            tree.heading(col, text=col)
            tree.column(col, width=140, anchor='center')

        vsb = ttk.Scrollbar(frame, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        vsb.pack(side='right', fill='y', pady=10)

        for i, linha in enumerate(linhas):
            tree.insert('', 'end', values=linha, tags=('impar',) if i % 2 else ())
        tree.tag_configure('impar', background='#FFF8FB')
        return tree

    def _df_para_linhas(self, df, ncols=6):
        return [tuple(str(v) if v is not None else '' for v in row[:ncols])
                for _, row in df.iterrows()]

    # ── telas ─────────────────────────────────────────────────────────────────

    def _tela_estoque(self):
        self._titulo('📦 Estoque de Mainha', 'Todos os produtos cadastrados')

        # barra de busca
        bf = ctk.CTkFrame(self.main, fg_color='transparent')
        bf.pack(fill='x', padx=30, pady=(10, 0))
        self._busca = ctk.CTkEntry(bf, placeholder_text='🔍  Buscar produto...',
                                   height=36, corner_radius=8,
                                   border_color='#DDDDDD', font=ctk.CTkFont(size=13))
        self._busca.pack(side='left', fill='x', expand=True)
        ctk.CTkButton(bf, text='Buscar', width=90, height=36,
                      fg_color=COR_ROSA, hover_color=COR_ROSA2, corner_radius=8,
                      command=self._filtrar).pack(side='left', padx=(8, 0))

        try:
            # usa visualizar_planilha() do back-end para carregar o DataFrame
            self._df = backend.visualizar_planilha()
            cols = list(self._df.columns[:6])
            self._tree = self._tabela(cols, self._df_para_linhas(self._df))
        except Exception as e:
            self._aviso(self.main, f'Erro ao carregar planilha: {e}', 'red')

    def _filtrar(self):
        termo = self._busca.get().lower().strip()
        df = self._df
        if termo:
            df = df[df.iloc[:, backend.COLUNA_PRODUTO]
                    .astype(str).str.lower().str.contains(termo)]
        for item in self._tree.get_children():
            self._tree.delete(item)
        for linha in self._df_para_linhas(df):
            self._tree.insert('', 'end', values=linha)

    # ──────────────────────────────────────────────────────────────────────────

    def _tela_venda(self):
        self._titulo('🛒 Registrar Venda', 'Informe o produto e a quantidade vendida')
        card = ctk.CTkFrame(self.main, fg_color=COR_CARD, corner_radius=12)
        card.pack(padx=30, pady=20, fill='x')
        inner = ctk.CTkFrame(card, fg_color='transparent')
        inner.pack(padx=30, pady=20, fill='x')

        self._vn = self._campo(inner, 'Nome do Produto', 'Ex: Perfume Natura Raiz')
        self._vq = self._campo(inner, 'Quantidade Vendida', 'Ex: 2')
        self._va = ctk.CTkLabel(inner, text='', font=ctk.CTkFont(size=12))
        self._va.pack(pady=(6, 0))
        self._botao(inner, '✅  Confirmar Venda', self._confirmar_venda).pack(fill='x', pady=(16, 0))

    def _confirmar_venda(self):
        nome = self._vn.get().strip()
        qtd  = self._vq.get().strip()
        if not nome or not qtd:
            self._va.configure(text='⚠ Preencha todos os campos.', text_color=COR_ALERTA)
            return
        try:
            qtd = int(qtd)
        except ValueError:
            self._va.configure(text='⚠ Quantidade deve ser um número inteiro.', text_color='red')
            return
        ok, msg = _registrar_venda(nome, qtd)
        self._va.configure(text=msg, text_color=COR_VERDE if ok else 'red')
        if ok:
            self._vn.delete(0, 'end')
            self._vq.delete(0, 'end')

    # ──────────────────────────────────────────────────────────────────────────

    def _tela_adicionar(self):
        self._titulo('➕ Novo Produto', 'Cadastre um novo item no estoque')
        card = ctk.CTkFrame(self.main, fg_color=COR_CARD, corner_radius=12)
        card.pack(padx=30, pady=20, fill='x')
        inner = ctk.CTkFrame(card, fg_color='transparent')
        inner.pack(padx=30, pady=20, fill='x')

        self._an = self._campo(inner, 'Nome do Produto',            'Ex: Creme Avon Care')
        self._aq = self._campo(inner, 'Quantidade em Estoque',      'Ex: 10')
        self._av = self._campo(inner, 'Já Vendidos (0 se novo)',    'Ex: 0')
        self._ap = self._campo(inner, 'Preço Unitário (R$)',        'Ex: 39.90')
        self._aa = ctk.CTkLabel(inner, text='', font=ctk.CTkFont(size=12))
        self._aa.pack(pady=(6, 0))
        self._botao(inner, '💾  Salvar Produto', self._salvar_produto).pack(fill='x', pady=(16, 0))

    def _salvar_produto(self):
        nome    = self._an.get().strip()
        qtd     = self._aq.get().strip()
        vendida = self._av.get().strip()
        preco   = self._ap.get().strip().replace(',', '.')
        if not all([nome, qtd, vendida, preco]):
            self._aa.configure(text='⚠ Preencha todos os campos.', text_color=COR_ALERTA)
            return
        try:
            _adicionar_produto(nome, int(qtd), int(vendida), float(preco))
            self._aa.configure(text=f'✅ "{nome}" adicionado!', text_color=COR_VERDE)
            for e in [self._an, self._aq, self._av, self._ap]:
                e.delete(0, 'end')
        except ValueError:
            self._aa.configure(text='⚠ Quantidade/Preço inválidos.', text_color='red')
        except Exception as ex:
            self._aa.configure(text=f'Erro: {ex}', text_color='red')

    # ──────────────────────────────────────────────────────────────────────────

    def _tela_repor(self):
        self._titulo('🔄 Repor Estoque', 'Adicione unidades a um produto já existente')
        card = ctk.CTkFrame(self.main, fg_color=COR_CARD, corner_radius=12)
        card.pack(padx=30, pady=20, fill='x')
        inner = ctk.CTkFrame(card, fg_color='transparent')
        inner.pack(padx=30, pady=20, fill='x')

        self._rn = self._campo(inner, 'Nome do Produto',       'Ex: Perfume Natura Raiz')
        self._rq = self._campo(inner, 'Quantidade que Chegou', 'Ex: 5')
        self._ra = ctk.CTkLabel(inner, text='', font=ctk.CTkFont(size=12))
        self._ra.pack(pady=(6, 0))
        self._botao(inner, '📥  Repor Estoque', self._confirmar_reposicao).pack(fill='x', pady=(16, 0))

    def _confirmar_reposicao(self):
        nome = self._rn.get().strip()
        qtd  = self._rq.get().strip()
        if not nome or not qtd:
            self._ra.configure(text='⚠ Preencha todos os campos.', text_color=COR_ALERTA)
            return
        try:
            qtd = int(qtd)
        except ValueError:
            self._ra.configure(text='⚠ Quantidade deve ser um número inteiro.', text_color='red')
            return
        ok, msg = _atualizar_estoque(nome, qtd)
        self._ra.configure(text=msg, text_color=COR_VERDE if ok else 'red')
        if ok:
            self._rn.delete(0, 'end')
            self._rq.delete(0, 'end')

    # ──────────────────────────────────────────────────────────────────────────

    def _tela_relatorio(self):
        self._titulo('📊 Relatório Geral', 'Resumo de faturamento e vendas')

        try:
            # usa visualizar_planilha() do back-end para os dados
            df = backend.visualizar_planilha()
        except Exception as e:
            self._aviso(self.main, f'Erro: {e}', 'red')
            return

        total    = pd.to_numeric(df.iloc[:, backend.COLUNA_TOTAL],   errors='coerce').sum()
        vendidas = pd.to_numeric(df.iloc[:, backend.COLUNA_VENDIDA], errors='coerce').sum()

        # cards de resumo
        cf = ctk.CTkFrame(self.main, fg_color='transparent')
        cf.pack(fill='x', padx=30, pady=20)
        for titulo, valor in [
            ('💰 Faturamento Total', f'R$ {total:,.2f}'),
            ('📦 Itens Vendidos',    f'{int(vendidas)} unidades'),
        ]:
            c = ctk.CTkFrame(cf, fg_color=COR_ROSA, corner_radius=14, width=200, height=90)
            c.pack(side='left', padx=(0, 16))
            c.pack_propagate(False)
            ctk.CTkLabel(c, text=titulo, font=ctk.CTkFont(size=12), text_color='white').pack(pady=(18, 2))
            ctk.CTkLabel(c, text=valor,  font=ctk.CTkFont(size=20, weight='bold'), text_color='white').pack()

        # top 10 mais vendidos
        ctk.CTkLabel(self.main, text='🏆 Produtos Mais Vendidos',
                     font=ctk.CTkFont(size=15, weight='bold'),
                     text_color=COR_TEXTO).pack(anchor='w', padx=30, pady=(10, 0))
        try:
            df['_v'] = pd.to_numeric(df.iloc[:, backend.COLUNA_VENDIDA], errors='coerce')
            top = df.dropna(subset=['_v']).sort_values('_v', ascending=False).head(10)
            self._tabela(
                ['Produto', 'Qtd Vendida', 'Faturado'],
                [(row.iloc[backend.COLUNA_PRODUTO],
                  int(row.iloc[backend.COLUNA_VENDIDA]),
                  f"R$ {float(row.iloc[backend.COLUNA_TOTAL] or 0):,.2f}")
                 for _, row in top.iterrows()]
            )
        except Exception as e:
            self._aviso(self.main, f'Erro no ranking: {e}', 'red')

    # ──────────────────────────────────────────────────────────────────────────

    def _tela_historico(self):
        self._titulo('📋 Histórico de Vendas', 'Todas as vendas registradas, da mais recente')
        try:
            # usa exibir_historico() indiretamente — lê a mesma aba 'Historico'
            df = pd.read_excel(backend.ARQUIVO_EXCEL, sheet_name='Historico')
            if df.empty:
                raise ValueError('vazio')
            df = df.iloc[::-1]  # mais recente primeiro, igual ao back-end
            self._tabela(list(df.columns),
                         [tuple(str(v) if v is not None else '' for v in row)
                          for _, row in df.iterrows()])
        except Exception:
            ctk.CTkLabel(self.main,
                         text='Nenhuma venda registrada ainda.',
                         font=ctk.CTkFont(size=14), text_color=COR_CINZA).pack(pady=40)

    # ──────────────────────────────────────────────────────────────────────────

    def _tela_baixo(self):
        self._titulo('⚠️ Estoque Baixo', 'Produtos com 1 ou menos unidades disponíveis')
        try:
            # usa produtos_faltando() do back-end para obter o DataFrame filtrado
            df = backend.visualizar_planilha()
            disp     = pd.to_numeric(df.iloc[:, backend.COLUNA_DISPONIVEL], errors='coerce')
            faltando = df[disp <= 1]

            if faltando.empty:
                ctk.CTkLabel(self.main,
                             text='✅  Tudo certo! Nenhum produto com estoque baixo.',
                             font=ctk.CTkFont(size=15), text_color=COR_VERDE).pack(pady=40)
                return
            self._tabela(list(faltando.columns[:6]), self._df_para_linhas(faltando))
        except Exception as e:
            self._aviso(self.main, f'Erro: {e}', 'red')


# ── entry point ───────────────────────────────────────────────────────────────
if __name__ == '__main__':
    App().mainloop()
