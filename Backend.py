from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
import matplotlib.pyplot as plt

ARQUIVO_EXCEL = 'AvonNatura.xlsx'
ABA_ESTOQUE = 'Estoque'

#indices de colunas
COLUNA_PRODUTO = 0
COLUNA_QTD = 1
COLUNA_VENDIDA = 2
COLUNA_DISPONIVEL = 3
COLUNA_PRECO = 4
COLUNA_TOTAL = 5

def visualizar_planilha():
    ler_planilha = load_workbook(ARQUIVO_EXCEL, data_only=True)
    sheet = ler_planilha[ABA_ESTOQUE]

    dados = []

    for linha in sheet.iter_rows(values_only=True):
        if all(celula is None for celula in linha):
            continue
        dados.append(linha)

    cabecalho = dados[0]
    dados_tabela = dados[1:]

    df = pd.DataFrame(dados_tabela, columns=cabecalho) #O pandas cria 0 1 2 3 4

    pd.set_option('display.max_columns', None)
    print(df.to_string(index=False)) #index false pra ignorar a coluna com as posições dos dados
    #to_string transforma tudo em string para aparecer bem clean
    return df
    #df agora é uma tabela pandas

def adicionar_produto():

    addPlanilha = load_workbook(ARQUIVO_EXCEL)
    sheet = addPlanilha[ABA_ESTOQUE]

    nome = input('Nome do produto: ')
    qtdProduto = int(input('Quantidade do produto: '))
    qtdVendida = int(input('Quantidade vendida até agora: '))
    valor = float(input('Valor do produto: '))

    linha = sheet.max_row
    while sheet.cell(row=linha, column=1).value is None:
        linha -= 1
    linha += 1

    dados = [
        nome,
        qtdProduto,
        qtdVendida,
        None,
        valor,
        None
    ]

    for coluna, dado in enumerate(dados, start=1):
        sheet.cell(row=linha, column=coluna, value=dado)

    sheet[f'D{linha}'] = f'=B{linha}-C{linha}'
    sheet[f'F{linha}'] = f'=C{linha}*E{linha}'

    addPlanilha.save(ARQUIVO_EXCEL)
    addPlanilha.close()

def aplicar_formulas():
    #usar para automaticamente adicionar as formulas nas colunas certas

    addPlanilha = load_workbook(ARQUIVO_EXCEL)

    sheet = addPlanilha[ABA_ESTOQUE]

    for linha in range(2, sheet.max_row + 1):

        produto = sheet[f'A{linha}'].value #pega o produto da coluna A

        if produto is None:
            continue

        if str(produto).strip().lower() == 'produto': #ignora cabeçalhos
            continue

        sheet[f'D{linha}'] = f'=B{linha}-C{linha}' #formula coluna D

        sheet[f'F{linha}'] = f'=C{linha}*E{linha}' #formula coluna F

    addPlanilha.save(ARQUIVO_EXCEL)
    addPlanilha.close()


def registrar_venda():

    addPlanilha = load_workbook(ARQUIVO_EXCEL)
    sheet = addPlanilha[ABA_ESTOQUE]

    nome = input('Nome do produto vendido: ')
    qtdVendida_input = int(input('Quantidade vendida: '))
    produto_encontrado = False

    for linha in range(2, sheet.max_row + 1):
        produto_celula = sheet[f'A{linha}'].value

        if produto_celula is None:
            continue

        if str(produto_celula).strip().lower() == nome.lower(): #pra n dar erro em caso de maiusculo/minusculo
            produto_encontrado = True

            # calculo do disponível direto das colunas B e C para evitar erro de string/formula
            qtd_total = sheet[f'B{linha}'].value or 0
            qtd_ja_vendida = sheet[f'C{linha}'].value or 0
            valor_unitario = sheet[f'E{linha}'].value or 0

            disponivel = qtd_total - qtd_ja_vendida

            if qtdVendida_input > disponivel:
                print(f'Estoque insuficiente. Disponível: {disponivel}')
                break

            #atualiza o estoque
            sheet[f'C{linha}'] = qtd_ja_vendida + qtdVendida_input

            #chama a função de histórico antes de fechar
            historico_vendas(addPlanilha, produto_celula, qtdVendida_input, valor_unitario)
            addPlanilha.save(ARQUIVO_EXCEL)
            print(f'Venda de {qtdVendida_input} unidades de "{produto_celula}" registrada!')
            break

    if not produto_encontrado:
        print('Produto não encontrado.')

    addPlanilha.close()

def buscar_produto():
    df = visualizar_planilha()

    nome = input('Nome do produto: ').lower()

    produtos = (df['Produto'].fillna('').astype(str))

    resultado = df[
        produtos.str.lower().str.contains(nome)
        ]

    if resultado.empty:
        print('\n❌ Nenhum produto encontrado.')
        return

    print('\n✅ Produtos encontrados:\n')
    print(resultado.to_string(index=False))

def produtos_faltando():
    df = visualizar_planilha()

    disponivel = pd.to_numeric( #converte para numero
        df.iloc[:, COLUNA_DISPONIVEL], # : = todas as linhas, sendo [: todas as linhas , da coluna "disponivel"]
        # iloc = index location
        errors='coerce'
        #se encontrar texto inválido: “valor inválido/vazio” (para n ter problemas com celulas da coluna que são texto
    )
    faltando = df[
        disponivel <= 1
     ]

    if faltando.empty:
        print('\n✅ Nenhum produto com estoque baixo.')
        return
    print('\n⚠ PRODUTOS COM ESTOQUE BAIXO:\n')
    for _, linha in faltando.iterrows(): #percorre linha por linha do DataFrame
        produto = linha.iloc[COLUNA_PRODUTO]
        qtd = linha.iloc[COLUNA_DISPONIVEL]
        print(f'⚠ {produto} → Apenas {qtd} unidades restantes')

def historico_vendas(addPlanilha, produto, quantidade, valor): #já é lançado conforme há registro de vendas

    if 'Historico' not in addPlanilha.sheetnames:
        addPlanilha.create_sheet('Historico')
        sheetHistorico = addPlanilha['Historico']

        sheetHistorico.append(["Data/Hora", "Produto", "Quantidade", "Valor Unit.", "Total"])
    else:
        sheetHistorico = addPlanilha['Historico']

    data_venda = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    total_venda = quantidade * valor

    sheetHistorico.append([data_venda, produto, quantidade, valor, total_venda])

def exibir_historico():
    sheet = pd.read_excel(ARQUIVO_EXCEL, sheet_name='Historico')
    if sheet.empty:
        print("\n[!] O histórico ainda está vazio.")
        return

    historico_recente = sheet.iloc[::-1] #inverte a ordem pros mais recentes aparecerem primeiro
    print(historico_recente.to_string(index=False))

def exportar_excel():
    pass
def atualizar_estoque(): #para quando receber uma nova remessa de produtos já existentes
    addPlanilha = load_workbook(ARQUIVO_EXCEL)
    sheet = addPlanilha[ABA_ESTOQUE]

    nome = input('Nome do produto para repor: ').lower()
    qtdNova = int(input('Quantidade que chegou: '))
    encontrado = False

    for linha in range(2, sheet.max_row + 1):
        if sheet[f'A{linha}'].value and str(sheet[f'A{linha}'].value).strip().lower() == nome:
            qtdAtual = sheet[f'B{linha}'].value or 0
            sheet[f'B{linha}'] = qtdAtual + qtdNova
            encontrado = True
            break

    if not encontrado:
        print('Produto não encontrado. Considere utilizar "Adicionar Produto"')
    addPlanilha.save(ARQUIVO_EXCEL)
    addPlanilha.close()

def ver_relatorio():  # relatorio de itens vendidos e faturamento total

    df = visualizar_planilha()

    total = pd.to_numeric(
        df.iloc[:, COLUNA_TOTAL],
        errors='coerce'
    )

    vendida = pd.to_numeric(
        df.iloc[:, COLUNA_VENDIDA],
        errors='coerce'
    )

    faturamento = total.sum()

    itens_vendidos = vendida.sum()

    print(f"Faturamento Total: R$ {faturamento:.2f}")

    print(f"Total de Itens Vendidos: {itens_vendidos:.0f} unidades")

    # graficos

    df_grafico = df.dropna(
        subset=[df.columns[COLUNA_PRODUTO]]
    )

    # remove linhas com texto inválido

    df_grafico = df_grafico[
        df_grafico.iloc[:, COLUNA_VENDIDA]
        .astype(str)
        .str.replace('.', '', regex=False)
        .str.isnumeric()
    ]

    # 10 mais vendidos

    top10 = df_grafico.sort_values(
        by=df.columns[COLUNA_VENDIDA],
        ascending=False
    ).head(10)

    plt.figure(figsize=(12, 6))

    plt.barh(
        top10.iloc[:, COLUNA_PRODUTO],
        top10.iloc[:, COLUNA_VENDIDA]
    )

    plt.title('TOP 10 Produtos Mais Vendidos')

    plt.xlabel('Quantidade Vendida')

    plt.ylabel('Produtos')

    plt.tight_layout()

    plt.show()

    # pizza

    total = pd.to_numeric(
        df_grafico.iloc[:, COLUNA_TOTAL],
        errors='coerce'
    )

    df_grafico = df_grafico[total.notna()]

    top5_faturamento = df_grafico.sort_values(
        by=df.columns[COLUNA_TOTAL],
        ascending=False
    ).head(5)

    plt.figure(figsize=(8, 8))

    plt.pie(
        top5_faturamento.iloc[:, COLUNA_TOTAL],
        labels=top5_faturamento.iloc[:, COLUNA_PRODUTO],
        autopct='%1.1f%%'
    )

    plt.title('5 maiores participações no Faturamento')

    plt.tight_layout()

    plt.show()

if __name__ == "__main__":

    '''visualizar_planilha()
    adicionar_produto()
    registrar_venda()
    produtos_faltando()
    ver_relatorio()
    exibir_historico()
    atualizar_estoque()'''
